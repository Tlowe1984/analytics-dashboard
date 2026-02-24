#!/usr/bin/env python3
"""
Sync milestones from Wearable Program Milestones SOT spreadsheet
Uses rclone to access Google Sheets and parse milestone data
"""

import sys
import json
import subprocess
import tempfile
import csv
from datetime import datetime, timedelta
import re

# Spreadsheet file in Google Drive
FILE_NAME = "Wearable Program Milestones SOT - For AI ／ User Consumption.xlsx"
RCLONE_CONFIG = "/home/ubuntu/.gdrive-rclone.ini"

def download_sheet_via_rclone():
    """Download Excel file from Google Drive using rclone"""
    try:
        # Download the Excel file from Google Drive
        result = subprocess.run(
            [
                "rclone", "copy",
                f"manus_google_drive:{FILE_NAME}",
                "/tmp/",
                "--config", RCLONE_CONFIG
            ],
            capture_output=True,
            text=True,
            timeout=120
        )
        
        if result.returncode != 0:
            print(f"rclone error: {result.stderr}", file=sys.stderr)
            return None
        
        return f"/tmp/{FILE_NAME}"
    except Exception as e:
        print(f"Error downloading file: {e}", file=sys.stderr)
        return None

def parse_week_to_date(week_str):
    """Convert week string like 'W5 2026' or date string to ISO date"""
    if not week_str or not isinstance(week_str, str):
        return None
    
    week_str = week_str.strip()
    
    # Match patterns like "W5 2026", "W5 '26", "W05 2026"
    week_match = re.match(r"W'?(\d+)\s*['\"]?(\d{2,4})?", week_str, re.IGNORECASE)
    if week_match:
        week_num = int(week_match.group(1))
        year_str = week_match.group(2)
        
        if year_str:
            year = int(year_str)
            # Handle 2-digit years
            if year < 100:
                year = 2000 + year
        else:
            year = 2026  # Default year
        
        # Convert ISO week to date (Monday of that week)
        # Week 1 is the week with Jan 4 in it
        jan4 = datetime(year, 1, 4)
        week1_monday = jan4 - timedelta(days=jan4.weekday())
        target_date = week1_monday + timedelta(weeks=week_num - 1)
        
        return target_date.strftime("%Y-%m-%d")
    
    # Try parsing as date in various formats
    for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%d-%b-%y"]:
        try:
            date_obj = datetime.strptime(week_str, fmt)
            return date_obj.strftime("%Y-%m-%d")
        except:
            continue
    
    return None

def categorize_milestone(product, milestone_name, type_hint=""):
    """Determine milestone type based on name and context"""
    name_lower = milestone_name.lower()
    type_lower = type_hint.lower() if type_hint else ""
    product_lower = product.lower()
    
    # Release/Launch dates
    if any(keyword in name_lower for keyword in ["launch", "release", "osd", "in market", "ship"]):
        return "release_milestones"
    
    # PDP Gates (Product Development Process)
    if any(keyword in name_lower for keyword in ["commit", "gate", "checkpoint", "concept", "discover", "define", "develop", "deliver"]):
        return "pdp_gates"
    
    # Software milestones
    if any(keyword in name_lower for keyword in ["fatp", "beta", "alpha", "zbb", "gmc", "fmc", "lbu-os", "dogfood"]):
        return "sw_milestones"
    
    # Hardware dates
    if any(keyword in name_lower for keyword in ["evt", "dvt", "pvt", "p1", "p2", "p3", "smt", "build"]):
        return "hw_dates"
    
    # Check type hint first (most explicit)
    if "gtm" in type_lower or "go-to-market" in type_lower or "go to market" in type_lower:
        return "gtm_milestones"
    if "sw" in type_lower or "software" in type_lower:
        return "sw_milestones"
    if "hw" in type_lower or "hardware" in type_lower:
        return "hw_dates"
    if "pdp" in type_lower or "gate" in type_lower:
        return "pdp_gates"
    
    # Check milestone name for GTM keywords
    if any(keyword in name_lower for keyword in ["gtm", "go-to-market", "go to market"]):
        return "gtm_milestones"
    
    # Default to PDP gates for ambiguous cases
    return "pdp_gates"

def parse_milestones_excel(excel_file):
    """Parse Excel file and extract milestones"""
    milestones = []
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb['Device_Milestones']
        
        # Find header row (row 2 based on earlier inspection)
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row and any(cell and 'product' in str(cell).lower() for cell in row):
                header_row = row_idx
                break
        
        if not header_row:
            print("Could not find header row", file=sys.stderr)
            return []
        
        # Get column indices
        headers = [cell for cell in ws[header_row]]
        col_indices = {}
        for idx, cell in enumerate(headers):
            if cell and cell.value:
                col_name = str(cell.value).strip().lower()
                if 'product' in col_name:
                    col_indices['product'] = idx
                elif 'milestone' in col_name and 'type' not in col_name:
                    col_indices['milestone'] = idx
                elif 'date' in col_name:
                    col_indices['date'] = idx
                elif 'type' in col_name:
                    col_indices['type'] = idx
        
        # Parse data rows
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not any(row):
                continue
            
            # Extract values
            product = str(row[col_indices.get('product', 0)] or '').strip()
            milestone_name = str(row[col_indices.get('milestone', 1)] or '').strip()
            date_value = row[col_indices.get('date', 2)]
            type_hint = str(row[col_indices.get('type', 3)] or '').strip()
            
            # Skip if missing essential fields
            if not product or not milestone_name or not date_value:
                continue
            
            # Parse date (could be datetime object or string)
            if isinstance(date_value, datetime):
                milestone_date = date_value.strftime("%Y-%m-%d")
            elif isinstance(date_value, (int, float)):
                # Excel serial date
                try:
                    excel_epoch = datetime(1899, 12, 30)
                    milestone_date = (excel_epoch + timedelta(days=date_value)).strftime("%Y-%m-%d")
                except:
                    continue
            else:
                milestone_date = parse_week_to_date(str(date_value))
                if not milestone_date:
                    continue
            
            # Categorize milestone
            milestone_type = categorize_milestone(product, milestone_name, type_hint)
            
            milestones.append({
                'product': product,
                'milestone_name': milestone_name,
                'milestone_date': milestone_date,
                'milestone_type': milestone_type,
                'original_type': type_hint if type_hint else None
            })
    
    except Exception as e:
        print(f"Error parsing Excel: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return []
    
    return milestones

def main():
    print("Downloading Wearable Program Milestones SOT spreadsheet...", file=sys.stderr)
    
    # Download spreadsheet
    excel_file = download_sheet_via_rclone()
    if not excel_file:
        print("Failed to download spreadsheet", file=sys.stderr)
        print("[]")
        return 1
    
    # Check if file exists
    import os
    if not os.path.exists(excel_file):
        print(f"Downloaded file not found: {excel_file}", file=sys.stderr)
        print("[]")
        return 1
    
    print(f"Parsing milestones from Excel file...", file=sys.stderr)
    
    # Parse milestones
    milestones = parse_milestones_excel(excel_file)
    
    print(f"Found {len(milestones)} milestones", file=sys.stderr)
    
    # Output as JSON
    print(json.dumps(milestones, indent=2))
    
    return 0

if __name__ == "__main__":
    sys.exit(main())
