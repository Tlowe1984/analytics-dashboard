#!/usr/bin/env python3
"""
Parser for upcoming reviews from three sign-up sheets:
- 2026 Wearables Reviews Sign-Up Sheet
- 2026 Product Reviews Sign-Up Sheet (tab 1)
- Systems Reviews Sign-Up Sheet

Extracts reviews for the next 14 days with:
- Review Type
- Week of review (with date in parenthesis)
- Topic
- Description
- Owner
"""

import openpyxl
import json
import sys
from datetime import datetime, timedelta
from pathlib import Path

def get_week_string(date_obj):
    """Convert date to 'W{week_num} ({Mon DD})' format"""
    week_num = date_obj.isocalendar()[1]
    date_str = date_obj.strftime("%b %d")
    return f"W{week_num} ({date_str})"

def parse_wearables_reviews(filepath):
    """Parse 2026 Wearables Reviews Sign-Up Sheet"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Wearables Reviews']
    
    reviews = []
    now = datetime.now()
    two_weeks = now + timedelta(days=14)
    
    # Header is in row 4
    for row_idx in range(5, ws.max_row + 1):
        first_date = ws.cell(row_idx, 1).value  # Column A: First Date for Review
        review_type = ws.cell(row_idx, 6).value  # Column F: Type of Review
        program = ws.cell(row_idx, 7).value  # Column G: Associated Program
        owner = ws.cell(row_idx, 10).value  # Column J: Driver/Requester
        title = ws.cell(row_idx, 11).value  # Column K: Review Title
        topic_summary = ws.cell(row_idx, 12).value  # Column L: Topic Summary
        scheduled_date = ws.cell(row_idx, 4).value  # Column D: Scheduled Date
        
        # Use scheduled date if available, otherwise first date
        review_date = scheduled_date if scheduled_date else first_date
        
        if isinstance(review_date, datetime) and now <= review_date <= two_weeks:
            # Use title for topic, program as fallback
            topic = title if title else program if program else "TBD"
            # Use topic summary as description if available
            description = topic_summary if topic_summary else title if title else program if program else "TBD"
            
            reviews.append({
                'review_type': 'Wearables Review',
                'week': get_week_string(review_date),
                'date': review_date.isoformat(),
                'topic': topic,
                'description': description,
                'owner': owner if owner else "TBD"
            })
    
    return reviews

def parse_product_reviews(filepath):
    """Parse 2026 Product Reviews Sign-Up Sheet (tab 1)"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Wearables Product Reviews Start']
    
    reviews = []
    now = datetime.now()
    two_weeks = now + timedelta(days=14)
    
    # Header is in row 2
    for row_idx in range(3, ws.max_row + 1):
        review_date = ws.cell(row_idx, 1).value  # Column A: Review Date
        pillar = ws.cell(row_idx, 2).value  # Column B: Pillar
        sponsor = ws.cell(row_idx, 3).value  # Column C: Sponsor
        presenter = ws.cell(row_idx, 7).value  # Column G: Presenter(s)
        review_type = ws.cell(row_idx, 8).value  # Column H: Review Type
        device = ws.cell(row_idx, 10).value  # Column J: Device
        title = ws.cell(row_idx, 11).value  # Column K: Review Title
        topic_summary = ws.cell(row_idx, 12).value  # Column L: Topic Summary
        
        if isinstance(review_date, datetime) and now <= review_date <= two_weeks:
            # Use title for topic, device/pillar as fallback
            topic = title if title else device if device else pillar if pillar else "TBD"
            # Use topic summary as description if available
            description = topic_summary if topic_summary else title if title else device if device else pillar if pillar else "TBD"
            owner = presenter if presenter else sponsor if sponsor else "TBD"
            
            reviews.append({
                'review_type': 'Product Review',
                'week': get_week_string(review_date),
                'date': review_date.isoformat(),
                'topic': topic,
                'description': description,
                'owner': owner
            })
    
    return reviews

def parse_systems_reviews(filepath):
    """Parse Systems Reviews Sign-Up Sheet"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['2026 Systems Reviews']
    
    reviews = []
    now = datetime.now()
    two_weeks = now + timedelta(days=14)
    
    # Header is in row 4
    for row_idx in range(5, ws.max_row + 1):
        requested_date = ws.cell(row_idx, 1).value  # Column A: Requested Review Date
        scheduled_date = ws.cell(row_idx, 3).value  # Column C: Scheduled Date
        review_type = ws.cell(row_idx, 5).value  # Column E: Type of Review
        program = ws.cell(row_idx, 6).value  # Column F: Associated Program
        owner = ws.cell(row_idx, 9).value  # Column I: Driver/Requester
        title = ws.cell(row_idx, 10).value  # Column J: Review Title
        topic_summary = ws.cell(row_idx, 11).value  # Column K: Topic Summary
        
        # Use scheduled date if available, otherwise requested date
        review_date = scheduled_date if scheduled_date else requested_date
        
        if isinstance(review_date, datetime) and now <= review_date <= two_weeks:
            topic = title if title else program if program else "TBD"
            # Use topic summary as description if available, otherwise use title
            description = topic_summary if topic_summary else title if title else program if program else "TBD"
            
            reviews.append({
                'review_type': 'Systems Review',
                'week': get_week_string(review_date),
                'date': review_date.isoformat(),
                'topic': topic,
                'description': description,
                'owner': owner if owner else "TBD"
            })
    
    return reviews

def main():
    # Parse all three spreadsheets
    wearables = parse_wearables_reviews('/tmp/2026 Wearables Reviews Sign-Up Sheet .xlsx')
    product = parse_product_reviews('/tmp/2026 Product Reviews Sign-Up Sheet.xlsx')
    systems = parse_systems_reviews('/tmp/Systems Reviews Sign-Up Sheet .xlsx')
    
    # Combine and sort by date
    all_reviews = wearables + product + systems
    all_reviews.sort(key=lambda x: x['date'])
    
    # Output as JSON
    print(json.dumps(all_reviews, indent=2))
    
    # Also save to file
    output_file = '/tmp/upcoming_reviews_data.json'
    with open(output_file, 'w') as f:
        json.dump(all_reviews, f, indent=2)
    
    print(f"\n✅ Parsed {len(all_reviews)} upcoming reviews ({len(wearables)} Wearables, {len(product)} Product, {len(systems)} Systems)", file=sys.stderr)
    print(f"📁 Saved to {output_file}", file=sys.stderr)

if __name__ == '__main__':
    main()
