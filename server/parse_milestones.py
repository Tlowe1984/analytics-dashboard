#!/usr/bin/env python3
"""
Parse Wearable Program Milestones SOT spreadsheet to extract in-market release dates.
This script downloads the spreadsheet via rclone and extracts release milestones.
"""

import sys
import json
import subprocess
from datetime import datetime, timedelta
from pathlib import Path

def download_spreadsheet():
    """Download the milestones spreadsheet from Google Drive"""
    spreadsheet_id = "13a184kjaZxLOo0CjNsWpmQ8Z3lhRJRtd1MUtCxyQZs8"
    output_path = "/tmp/milestones.xlsx"
    
    # Try to download using rclone with the file by searching for it
    cmd = [
        "rclone", "copy",
        f"manus_google_drive:Wearable Program Milestones SOT - For AI / User Consumption",
        "/tmp/",
        "--config", "/home/ubuntu/.gdrive-rclone.ini",
        "--drive-export-formats", "xlsx",
        "--drive-skip-gdocs=false"
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    if Path(output_path).exists():
        return output_path
    
    # If that didn't work, the file might have a different name
    # List all xlsx files in /tmp that might be it
    xlsx_files = list(Path("/tmp").glob("*milestone*.xlsx"))
    if xlsx_files:
        return str(xlsx_files[0])
    
    xlsx_files = list(Path("/tmp").glob("Wearable*.xlsx"))
    if xlsx_files:
        return str(xlsx_files[0])
        
    raise FileNotFoundError(f"Could not download milestones spreadsheet. rclone output: {result.stderr}")

def parse_milestones(xlsx_path):
    """Parse the milestones spreadsheet and extract in-market release dates"""
    try:
        import openpyxl
    except ImportError:
        subprocess.run(["sudo", "pip3", "install", "-q", "openpyxl"], check=True)
        import openpyxl
    
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    
    milestones = []
    
    # Find the header row
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        row_str = ' '.join([str(cell).lower() if cell else '' for cell in row])
        if 'product' in row_str and 'date' in row_str:
            header_row = row_idx
            break
    
    if not header_row:
        print("Warning: Could not find header row", file=sys.stderr)
        return []
    
    # Get column indices
    headers = [str(cell).strip() if cell else '' for cell in ws[header_row]]
    
    product_col = None
    milestone_col = None
    date_col = None
    
    for idx, header in enumerate(headers):
        header_lower = header.lower()
        if 'product' in header_lower or 'program' in header_lower:
            product_col = idx
        elif 'milestone' in header_lower or 'gate' in header_lower or 'type' in header_lower:
            milestone_col = idx
        elif 'date' in header_lower or 'week' in header_lower:
            date_col = idx
    
    if product_col is None or date_col is None:
        print(f"Warning: Could not find required columns. Headers: {headers}", file=sys.stderr)
        return []
    
    # Parse data rows
    current_date = datetime.now()
    one_month_later = current_date + timedelta(days=30)
    
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or len(row) <= max(product_col, date_col):
            continue
        
        product = row[product_col]
        date_val = row[date_col]
        milestone = row[milestone_col] if milestone_col is not None and len(row) > milestone_col else None
        
        if not product or not date_val:
            continue
        
        # Skip if not an in-market release
        if milestone:
            milestone_str = str(milestone).lower()
            if 'launch' not in milestone_str and 'release' not in milestone_str and 'market' not in milestone_str:
                continue
        
        # Parse date
        release_date = None
        if isinstance(date_val, datetime):
            release_date = date_val
        elif isinstance(date_val, str):
            # Try to parse week format (e.g., "W16 (Apr 14)")
            if 'W' in date_val or 'w' in date_val:
                # Extract date from parentheses
                import re
                match = re.search(r'\(([^)]+)\)', date_val)
                if match:
                    date_str = match.group(1)
                    try:
                        # Try parsing "Apr 14" format
                        release_date = datetime.strptime(f"{date_str} 2026", "%b %d %Y")
                    except:
                        pass
        
        if not release_date:
            continue
        
        # Only include releases within the next month
        if current_date <= release_date <= one_month_later:
            milestones.append({
                "product": str(product).strip(),
                "milestone_type": str(milestone).strip() if milestone else "Launch",
                "date": release_date.strftime("%Y-%m-%d"),
                "week": f"W{release_date.isocalendar()[1]}",
                "display_date": release_date.strftime("%b %d")
            })
    
    # Sort by date
    milestones.sort(key=lambda x: x['date'])
    
    return milestones

def main():
    try:
        # Download spreadsheet
        xlsx_path = download_spreadsheet()
        
        # Parse milestones
        milestones = parse_milestones(xlsx_path)
        
        # Output as JSON
        print(json.dumps(milestones, indent=2))
        
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
