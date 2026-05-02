#!/usr/bin/env python3
"""
Parse Wearable Program Milestones SOT spreadsheet (XLSX format)
Handles computed values from formulas
"""

import sys
import json
import subprocess
from datetime import datetime, timedelta
import re

def download_spreadsheet():
    """Download spreadsheet using rclone - uses Aggregation Sheet which has actual data"""
    try:
        # Try the Aggregation Sheet first (has actual data in Release_Beast_Data and Smart_Sheet_Data)
        agg_file = "/tmp/Wearables Device Program Milestones - Aggregation Sheet.xlsx"
        subprocess.run(["rm", "-f", agg_file], check=False)
        
        result = subprocess.run(
            [
                "rclone", "copy",
                "--drive-shared-with-me",
                "manus_google_drive:Wearables Device Program Milestones - Aggregation Sheet.xlsx",
                "/tmp/",
                "--config", "/home/ubuntu/.gdrive-rclone.ini",
                "--ignore-times",
                "--no-check-certificate"
            ],
            capture_output=True,
            text=True,
            timeout=120
        )
        
        if result.returncode == 0:
            import os
            if os.path.exists(agg_file) and os.path.getsize(agg_file) > 10000:
                print(f"Using Aggregation Sheet: {agg_file}", file=sys.stderr)
                return agg_file
        
        # Fallback to SOT file
        sot_file = "/tmp/Wearable Program Milestones SOT - For AI ／ User Consumption.xlsx"
        subprocess.run(["rm", "-f", sot_file], check=False)
        
        result = subprocess.run(
            [
                "rclone", "copy",
                "manus_google_drive:Wearable Program Milestones SOT - For AI ／ User Consumption.xlsx",
                "/tmp/",
                "--config", "/home/ubuntu/.gdrive-rclone.ini",
                "--ignore-times",
                "--no-check-certificate"
            ],
            capture_output=True,
            text=True,
            timeout=120
        )
        
        if result.returncode != 0:
            print(f"rclone error: {result.stderr}", file=sys.stderr)
            return None
        
        return sot_file
    except Exception as e:
        print(f"Error downloading spreadsheet: {e}", file=sys.stderr)
        return None

def excel_date_to_iso(excel_date):
    """Convert Excel date number to ISO date string"""
    try:
        if isinstance(excel_date, (int, float)):
            # Excel dates are days since 1899-12-30
            base_date = datetime(1899, 12, 30)
            date_obj = base_date + timedelta(days=excel_date)
            return date_obj.strftime("%Y-%m-%d")
        elif isinstance(excel_date, datetime):
            return excel_date.strftime("%Y-%m-%d")
        elif isinstance(excel_date, str):
            # Try parsing as date
            for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y"]:
                try:
                    date_obj = datetime.strptime(excel_date, fmt)
                    return date_obj.strftime("%Y-%m-%d")
                except:
                    continue
        return None
    except:
        return None

def categorize_milestone(milestone_type_str):
    """Map milestone type string to enum"""
    if not milestone_type_str:
        return "sw_milestones"  # Default to SW
    
    type_lower = milestone_type_str.lower()
    
    # Only classify as PDP if explicitly labeled "PDP Milestones"
    if "pdp" in type_lower and "milestone" in type_lower:
        return "pdp_gates"
    # SDP (Software Development Program) milestones - separate type
    elif "sdp" in type_lower and "milestone" in type_lower:
        return "sdp_milestones"
    # GTM (Go-To-Market) milestones  
    elif "gtm" in type_lower or "go-to-market" in type_lower or "go to market" in type_lower:
        return "gtm_milestones"
    # Factory SW milestones
    elif "factory" in type_lower and "sw" in type_lower:
        return "sw_milestones"
    elif "sw" in type_lower or "software" in type_lower:
        return "sw_milestones"
    elif "hw" in type_lower or "hardware" in type_lower or "build" in type_lower:
        return "hw_dates"
    elif "silicon" in type_lower:
        return "hw_dates"
    elif "launch" in type_lower or "release" in type_lower:
        return "release_milestones"
    else:
        return "sw_milestones"  # Default to SW

def parse_xlsx(xlsx_file):
    """Parse XLSX file and extract milestones - handles both SOT and Aggregation Sheet formats"""
    try:
        import openpyxl
        import os
        
        wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        milestones = []
        
        # Check if this is the Aggregation Sheet (has Release_Beast_Data sheet)
        if 'Release_Beast_Data' in wb.sheetnames:
            print(f"Detected Aggregation Sheet format", file=sys.stderr)
            milestones.extend(parse_release_beast_data(wb))
            milestones.extend(parse_smart_sheet_data(wb))
            print(f"Total milestones from Aggregation Sheet: {len(milestones)}", file=sys.stderr)
            return milestones
        
        # Fall back to original SOT parsing (Consolidated View)
        ws = wb.active
        header_row = None
        
        # Find header row
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and any(cell and isinstance(cell, str) and 'product' in cell.lower() for cell in row):
                header_row = i
                headers = [str(cell).strip() if cell else '' for cell in row]
                break
        
        if not header_row:
            print("Could not find header row", file=sys.stderr)
            return []
        
        # Find column indices
        product_idx = next((i for i, h in enumerate(headers) if h and 'product' in h.lower()), None)
        milestone_idx = next((i for i, h in enumerate(headers) if h and 'milestone' in h.lower()), None)
        date_idx = next((i for i, h in enumerate(headers) if h and 'date' in h.lower()), None)
        type_idx = next((i for i, h in enumerate(headers) if h and 'type' in h.lower()), None)
        
        if product_idx is None or milestone_idx is None or date_idx is None:
            print(f"Could not find required columns. Headers: {headers}", file=sys.stderr)
            return []
        
        # Parse data rows
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or len(row) <= max(product_idx, milestone_idx, date_idx):
                continue
            
            product = row[product_idx]
            milestone_name = row[milestone_idx]
            date_value = row[date_idx]
            milestone_type_str = row[type_idx] if type_idx is not None and len(row) > type_idx else None
            
            if not product or not milestone_name or not date_value:
                continue
            
            product = str(product).strip()
            milestone_name = str(milestone_name).strip()
            
            milestone_date = excel_date_to_iso(date_value)
            if not milestone_date:
                continue
            
            milestone_type = categorize_milestone(milestone_type_str)
            
            milestones.append({
                'product': product,
                'milestone_name': milestone_name,
                'milestone_date': milestone_date,
                'milestone_type': milestone_type,
                'original_type': str(milestone_type_str) if milestone_type_str else None
            })
        
        return milestones
    
    except Exception as e:
        print(f"Error parsing XLSX: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return []

def parse_release_beast_data(wb):
    """Parse Release_Beast_Data sheet from Aggregation Sheet"""
    milestones = []
    try:
        ws = wb['Release_Beast_Data']
        headers = [str(cell.value).strip() if cell.value else '' for cell in list(ws.iter_rows(max_row=1))[0]]
        
        product_idx = next((i for i, h in enumerate(headers) if h and 'product_label' in h.lower()), None)
        if product_idx is None:
            product_idx = next((i for i, h in enumerate(headers) if h and h.lower() == 'product'), None)
        milestone_idx = next((i for i, h in enumerate(headers) if h and 'release_event' in h.lower()), None)
        date_idx = next((i for i, h in enumerate(headers) if h and 'current_planned_date' in h.lower()), None)
        if date_idx is None:
            date_idx = next((i for i, h in enumerate(headers) if h and 'initial_planned_date' in h.lower()), None)
        type_idx = next((i for i, h in enumerate(headers) if h and 'milestone type' in h.lower()), None)
        
        if product_idx is None or milestone_idx is None or date_idx is None:
            print(f"Release_Beast_Data: Could not find columns. Headers: {headers}", file=sys.stderr)
            return []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(product_idx, milestone_idx, date_idx):
                continue
            
            product = row[product_idx]
            milestone_name = row[milestone_idx]
            date_value = row[date_idx]
            milestone_type_str = row[type_idx] if type_idx is not None and len(row) > type_idx else None
            
            if not product or not milestone_name or not date_value:
                continue
            
            product = str(product).strip()
            milestone_name = str(milestone_name).strip()
            
            milestone_date = excel_date_to_iso(date_value)
            if not milestone_date:
                continue
            
            milestone_type = categorize_milestone(milestone_type_str or milestone_name)
            
            milestones.append({
                'product': product,
                'milestone_name': milestone_name,
                'milestone_date': milestone_date,
                'milestone_type': milestone_type,
                'original_type': str(milestone_type_str) if milestone_type_str else None
            })
        
        print(f"Release_Beast_Data: parsed {len(milestones)} milestones", file=sys.stderr)
    except Exception as e:
        print(f"Error parsing Release_Beast_Data: {e}", file=sys.stderr)
    return milestones

def parse_smart_sheet_data(wb):
    """Parse Smart_Sheet_Data sheet from Aggregation Sheet"""
    milestones = []
    try:
        ws = wb['Smart_Sheet_Data']
        headers = [str(cell.value).strip() if cell.value else '' for cell in list(ws.iter_rows(max_row=1))[0]]
        
        product_idx = next((i for i, h in enumerate(headers) if h and 'device_name' in h.lower()), None)
        milestone_idx = next((i for i, h in enumerate(headers) if h and 'milestone_name' in h.lower()), None)
        date_idx = next((i for i, h in enumerate(headers) if h and 'milestone_date' in h.lower()), None)
        
        if product_idx is None or milestone_idx is None or date_idx is None:
            print(f"Smart_Sheet_Data: Could not find columns. Headers: {headers}", file=sys.stderr)
            return []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(product_idx, milestone_idx, date_idx):
                continue
            
            product = row[product_idx]
            milestone_name = row[milestone_idx]
            date_value = row[date_idx]
            
            if not product or not milestone_name or not date_value:
                continue
            
            product = str(product).strip()
            milestone_name = str(milestone_name).strip()
            
            milestone_date = excel_date_to_iso(date_value)
            if not milestone_date:
                continue
            
            milestone_type = categorize_milestone(milestone_name)
            
            milestones.append({
                'product': product,
                'milestone_name': milestone_name,
                'milestone_date': milestone_date,
                'milestone_type': milestone_type,
                'original_type': None
            })
        
        print(f"Smart_Sheet_Data: parsed {len(milestones)} milestones", file=sys.stderr)
    except Exception as e:
        print(f"Error parsing Smart_Sheet_Data: {e}", file=sys.stderr)
    return milestones

def main():
    print("Downloading Wearable Program Milestones SOT spreadsheet...", file=sys.stderr)
    
    # Download spreadsheet
    xlsx_file = download_spreadsheet()
    if not xlsx_file:
        print("Failed to download spreadsheet", file=sys.stderr)
        print("[]")
        return 1
    
    print(f"Parsing milestones from {xlsx_file}...", file=sys.stderr)
    
    # Parse milestones
    milestones = parse_xlsx(xlsx_file)
    
    print(f"Found {len(milestones)} milestones", file=sys.stderr)
    
    # Output as JSON
    print(json.dumps(milestones, indent=2))
    
    return 0

if __name__ == "__main__":
    sys.exit(main())
